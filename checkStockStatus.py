import datetime

from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os.path
import time
from datetime import date, datetime

""""This method is used to check the stock status in Ozstock."""
def checkOzstock(url):
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'lxml')
    images = soup.findAll('img', {"src": True})
    for image in images:
        # Check whether the product is in stock or not
        if str(image['src']) == "/images/soldout_bt.gif":
            return "Out of stock"
            break

    else:
        return "In stock"


""""This method is used to check the stock status in Ebaystore."""
def checkEbay(url):
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'lxml')
    avalible = soup.find('span', {"id": "qtySubTxt"})
    if avalible:
        txt = avalible.findChild().string
        """Find the number from string"""
        stock = [int(s) for s in txt.split() if s.isdigit()]
    
        if stock[0] == 0:
            # Check whether the product is in stock or not
            return "Out of stock"
    
        else:
            return "In stock"

    else:
        return "Out of stock"

#This program is used to check if the product in Ozstock website is sold out.
"""url1 = input("Please enter the product's link from Ozstock:")
print(checkOzstock(url1))"""

""""This program is used to check if the product in Ozstock website is sold out.
url2 = input("Please enter the product's link from Ebay:")
print(checkEbay(url2))"""

""""This method is used to check if the date is one months ago."""
def checkDate(d):
    today = date.today()
    print(str(today-d))
    if(int(today-d)> 31):
        return True
    else:
        return False


""""This method is used to update the stock status by both online store."""
def updateStoreStatus():
    if os.path.exists('stockstatus.xlsx'):
        wb = load_workbook('stockstatus.xlsx')
        ws = wb.active

        for rowNumber in range(2, ws.max_row+1):
            if(checkOzstock(ws["B"+str(rowNumber)].value) == "Out of stock" or checkEbay(ws["C"+str(rowNumber)].value) == "Out of stock"):
                ws["D"+str(rowNumber)].value = "Out of stock"
            else:
                ws["D" + str(rowNumber)].value = "In stock"
                d = ws["F" + str(rowNumber)].value.date()
                if(checkDate(d)):
                    ws["F" + str(rowNumber)].fill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')


        wb.save('stockstatus.xlsx')

    else:
        print("The excel file is missing or not readable.")

if __name__ == '__main__':
    while True:
        updateStoreStatus()
        print("Updated once.")
        time.sleep(10.0)
