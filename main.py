import datetime

from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os.path
import sys
import time
from datetime import date, datetime, timedelta
import random

""""This method is used to check the stock status in Ozstock."""


def checkOzstock(url):
    if (url == None):
        return "In stock"
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'lxml')
    images = soup.findAll('img', {"src": True})
    for image in images:
        # Check whether the product is in stock or not
        if str(image['src']) == "/images/soldout_bt.gif":
            return "Out of stock"
            break

    else:
        return "In stock"


""""This method is used to check the stock status in Ebaystore."""


def checkEbay(url):
    if(url == None):
        return "In stock"
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'lxml')
    avalible = soup.find('span', {"id": "qtySubTxt"})
    oostext = soup.find('span', {"id": "w1-5-_msg"})
    if (avalible):
        txtava = avalible.findChild().string
        """Find the number from string"""

        stock = [int(s) for s in txtava.split() if s.isdigit()]

        if stock[0] == 0:
            # Check whether the product is in stock or not
            return "Out of stock"

    elif (oostext):
        txtoos = oostext.find("no longer available").string
        if (txtoos != None):
            return "Out of stock"

    else:
        return "In stock"


# This program is used to check if the product in Ozstock website is sold out.
"""url1 = input("Please enter the product's link from Ozstock:")
print(checkOzstock(url1))"""

""""This program is used to check if the product in Ozstock website is sold out.
url2 = input("Please enter the product's link from Ebay:")
print(checkEbay(url2))"""

""""This method is used to check if the date is one months ago."""


def checkDate(d):
    today = date.today()
    if (today > d + timedelta(days=30)):
        return True
    else:
        return False


""""This method is used to update the stock status by both online store."""


def updateStoreStatus():
    if os.path.exists('excel/stockstatus.xlsx'):
        wb = load_workbook('excel/stockstatus.xlsx')
        ws = wb.active

        for rowNumber in range(2, ws.max_row + 1):
            if (ws["G" + str(rowNumber)].value == None and ws["H" + str(rowNumber)].value == None):
                continue
            elif (ws["I" + str(rowNumber)].value != "Out of stock"):
                if (checkOzstock(ws["G" + str(rowNumber)].value) == "Out of stock" or checkEbay(
                        ws["H" + str(rowNumber)].value) == "Out of stock"):
                    ws["I" + str(rowNumber)].value = "Out of stock"
                    print(ws["A" + str(rowNumber)].value + " was just sold.")
                else:
                    ws["I" + str(rowNumber)].value = "In stock"

                    """highlight the posted items expired."""
                    if(ws["K" + str(rowNumber)].value == None):
                        continue
                    elif (checkDate(ws["K" + str(rowNumber)].value.date())):
                        ws["J" + str(rowNumber)].fill = PatternFill(start_color='FFFF0000',
                                                                    end_color='FFFF0000',
                                                                    fill_type='solid')
                        ws["K" + str(rowNumber)].fill = PatternFill(start_color='FFFF0000',
                                                                    end_color='FFFF0000',
                                                                    fill_type='solid')
                        print(ws["A" + str(rowNumber)].value + " product is expired.")
            else:
                continue

        wb.save('excel/stockstatus.xlsx')

    else:
        print("The excel file is missing or not readable.")


if __name__ == '__main__':
    print("The program started!")
    while True:
        updateStoreStatus()
        print("Updated on " + str(datetime.now()) + "\n\n\n")

        time.sleep(3000 + random.randint(1,600))
        """This program will update in random time from 0 to 5 minutes."""
